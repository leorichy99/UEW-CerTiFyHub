"""
Student record import service.

Accepts a CSV or XLSX file in-memory, validates each row, and creates
StudentRecord rows in a Draft batch. Supports field mapping so that
column names in the source file need not match system field names.
"""

import csv
import io
from datetime import date, datetime
from typing import Iterable, List, Dict, Any, Optional

from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone

from registry.models import (
    IssuanceBatch, StudentRecord, ImportBatch,
    Faculty, Department,
)
from registry.repositories import ImportBatchRepository
from registry.services.batch_lifecycle_service import BatchLifecycleService


SYSTEM_FIELDS = {
    'index_number': {'label': 'Index Number', 'required': True},
    'first_name': {'label': 'First Name', 'required': True},
    'middle_name': {'label': 'Middle Name', 'required': False},
    'last_name': {'label': 'Last Name', 'required': True},
    'gender': {'label': 'Gender', 'required': False},
    'institutional_email': {'label': 'Institutional Email', 'required': True},
    'programme': {'label': 'Programme', 'required': True},
    'class_of_degree': {'label': 'Class of Degree', 'required': True},
    'date_of_completion': {'label': 'Date of Completion', 'required': True},
    'date_of_admission': {'label': 'Date of Admission', 'required': False},
    'faculty': {'label': 'Faculty', 'required': False},
    'department': {'label': 'Department', 'required': False},
}

# Aliases used for auto-mapping (case-insensitive)
KNOWN_ALIASES = {
    'index_number': [
        'index number', 'index no', 'index no.', 'reg. no.', 'reg no',
        'registration number', 'student id', 'student number', 'id number',
    ],
    'first_name': [
        'first name', 'firstname', 'first', 'given name', 'student first name',
    ],
    'middle_name': [
        'middle name', 'middle names', 'other names', 'othername', 'other name',
        'other', 'student middle name', 'student other names',
    ],
    'last_name': [
        'last name', 'lastname', 'last', 'surname', 'family name',
        'student last name',
    ],
    'gender': ['gender', 'sex', 'student gender'],
    'institutional_email': [
        'institutional email', 'email', 'student email', 'e-mail',
        'uew email', 'institutional e-mail',
    ],
    'programme': [
        'programme', 'program', 'programme of study', 'course',
        'degree programme', 'program of study',
    ],
    'class_of_degree': [
        'class of degree', 'degree class', 'classification',
        'class', 'degree classification', 'award class',
    ],
    'date_of_completion': [
        'date of completion', 'completion date', 'date completed',
        'graduation date', 'date of graduation', 'award date',
    ],
    'date_of_admission': [
        'date of admission', 'admission date', 'date admitted',
        'enrolment date', 'enrollment date',
    ],
    'faculty': ['faculty', 'school', 'college'],
    'department': ['department', 'dept', 'dept.', 'unit'],
}

CLASS_OF_DEGREE_CHOICES = {
    'First Class', 'Second Class Upper', 'Second Class Lower', 'Third Class',
    'Pass', 'First', 'Second Upper', 'Second Lower', '2nd Upper', '2nd Lower',
    '1st Class', '2nd Class Upper', '2nd Class Lower', '3rd Class',
}


class ImportRejected(Exception):
    """Raised when the entire file is rejected (e.g. unreadable, missing headers)."""


class ImportService:
    def __init__(self, batch_repo=None, lifecycle=None):
        self.batch_repo = batch_repo or ImportBatchRepository()
        self.lifecycle = lifecycle or BatchLifecycleService()

    # ── Public API ───────────────────────────────────────────────────────

    @staticmethod
    def suggest_mapping(columns: List[str]) -> Dict[str, Optional[str]]:
        """Suggest a mapping from source columns to system fields using aliases."""
        mapping: Dict[str, Optional[str]] = {}
        used_columns = set()
        normalized_columns = {c: c.lower().strip() for c in columns}

        for field, aliases in KNOWN_ALIASES.items():
            match = None
            # Exact or alias match
            for col, norm in normalized_columns.items():
                if norm in aliases or norm == field.replace('_', ' '):
                    match = col
                    break
            if match and match not in used_columns:
                mapping[field] = match
                used_columns.add(match)
            else:
                mapping[field] = None

        return mapping

    def preview(self, *, file_name: str, raw_bytes: bytes, mapping: Dict[str, Optional[str]],
                batch: Optional[IssuanceBatch] = None, max_rows: int = 50):
        """Parse up to *max_rows* using *mapping* and return preview + issues."""
        rows = list(self._parse_rows_raw(file_name, raw_bytes))
        if not rows:
            raise ImportRejected('File is empty or unreadable.')

        total_rows = len(rows)
        preview_rows = []
        issues = []
        seen_indices = set()
        existing_indices = set()
        if batch:
            existing_indices = set(
                StudentRecord.objects
                .filter(batch=batch)
                .values_list('index_number', flat=True)
            )

        faculty_by_code = {f.code: f for f in Faculty.objects.all()}
        department_by_code = {d.code: d for d in Department.objects.all()}

        for line_no, raw in enumerate(rows[:max_rows], start=2):
            mapped = self._apply_mapping(raw, mapping)
            try:
                cleaned = self._validate_and_clean_row(
                    mapped, faculty_by_code, department_by_code,
                )
            except ValidationError as e:
                issues.append({
                    'row': line_no,
                    'field': getattr(e, 'field', '') or '',
                    'message': '; '.join(e.messages) if hasattr(e, 'messages') else str(e),
                })
                preview_rows.append({
                    **{k: mapped.get(k, '') for k in SYSTEM_FIELDS},
                    '_status': 'error',
                    '_issue': str(e),
                })
                continue

            idx = cleaned['index_number']
            if idx in seen_indices:
                issues.append({
                    'row': line_no, 'field': 'index_number',
                    'message': f"Duplicate index number {idx} in file.",
                })
                preview_rows.append({
                    **{k: mapped.get(k, '') for k in SYSTEM_FIELDS},
                    '_status': 'error',
                    '_issue': f"Duplicate index number {idx} in file.",
                })
                continue
            if idx in existing_indices:
                issues.append({
                    'row': line_no, 'field': 'index_number',
                    'message': f"Index number {idx} already exists in this batch.",
                })
                preview_rows.append({
                    **{k: mapped.get(k, '') for k in SYSTEM_FIELDS},
                    '_status': 'error',
                    '_issue': f"Index number {idx} already exists in this batch.",
                })
                continue
            seen_indices.add(idx)

            preview_rows.append({
                **{k: cleaned.get(k, '') for k in SYSTEM_FIELDS},
                '_status': 'valid',
                '_issue': '',
            })

        return {
            'preview_rows': preview_rows,
            'issues': issues,
            'total_rows': total_rows,
            'estimated_valid': total_rows - len(issues),  # rough estimate
            'estimated_issues': len(issues),
        }

    @transaction.atomic
    def process_async(self, *, import_batch_id: str, batch: IssuanceBatch,
                      uploaded_by, file_name: str, raw_bytes: bytes,
                      mapping: Dict[str, Optional[str]],
                      skip_invalid: bool = False,
                      progress_callback=None):
        """Parse full file with mapping, persist valid rows, update ImportBatch."""
        if not self.lifecycle.can_edit_records(batch):
            raise ImportRejected(
                'Student records can only be uploaded to a Draft batch.'
            )

        rows = list(self._parse_rows_raw(file_name, raw_bytes))
        if not rows:
            raise ImportRejected('File is empty or unreadable.')

        import_batch = self.batch_repo.get(import_batch_id)
        if not import_batch:
            raise ImportRejected('Import batch not found.')

        success = 0
        skipped = 0
        errors: List[Dict[str, Any]] = []

        faculty_by_code = {f.code: f for f in Faculty.objects.all()}
        department_by_code = {d.code: d for d in Department.objects.all()}

        seen_indices = set()
        existing_indices = set(
            StudentRecord.objects
            .filter(batch=batch)
            .values_list('index_number', flat=True)
        )

        total = len(rows)

        for line_no, raw in enumerate(rows, start=2):
            mapped = self._apply_mapping(raw, mapping)
            try:
                cleaned = self._validate_and_clean_row(
                    mapped, faculty_by_code, department_by_code,
                )
            except ValidationError as e:
                if not skip_invalid:
                    errors.append({
                        'row': line_no,
                        'field': getattr(e, 'field', '') or '',
                        'message': '; '.join(e.messages) if hasattr(e, 'messages') else str(e),
                    })
                else:
                    skipped += 1
                    errors.append({
                        'row': line_no,
                        'field': getattr(e, 'field', '') or '',
                        'message': '; '.join(e.messages) if hasattr(e, 'messages') else str(e),
                    })
                if progress_callback:
                    progress_callback(processed=line_no - 1, total=total,
                                      valid=success, skipped=skipped, errors=len(errors))
                continue

            idx = cleaned['index_number']
            if idx in seen_indices:
                if not skip_invalid:
                    errors.append({
                        'row': line_no, 'field': 'index_number',
                        'message': f"Duplicate index number {idx} in file.",
                    })
                else:
                    skipped += 1
                    errors.append({
                        'row': line_no, 'field': 'index_number',
                        'message': f"Duplicate index number {idx} in file.",
                    })
                if progress_callback:
                    progress_callback(processed=line_no - 1, total=total,
                                      valid=success, skipped=skipped, errors=len(errors))
                continue
            if idx in existing_indices:
                skipped += 1
                errors.append({
                    'row': line_no, 'field': 'index_number',
                    'message': f"Index number {idx} already exists in this batch.",
                })
                if progress_callback:
                    progress_callback(processed=line_no - 1, total=total,
                                      valid=success, skipped=skipped, errors=len(errors))
                continue
            seen_indices.add(idx)

            StudentRecord.objects.create(
                batch=batch, import_batch=import_batch, **cleaned,
            )
            success += 1

            if progress_callback and (line_no - 1) % 50 == 0:
                progress_callback(processed=line_no - 1, total=total,
                                  valid=success, skipped=skipped, errors=len(errors))

        if progress_callback:
            progress_callback(processed=total, total=total,
                              valid=success, skipped=skipped, errors=len(errors))

        import_batch_status = (
            ImportBatch.STATUS_COMPLETED if not errors
            else ImportBatch.STATUS_COMPLETED_WITH_ERRORS
        )
        ImportBatch.objects.filter(pk=import_batch.pk).update(
            total_rows=total,
            success_count=success,
            skipped_count=skipped,
            error_count=len(errors),
            error_log=errors,
            status=import_batch_status,
            completed_at=timezone.now(),
            mapping_configuration=mapping,
        )
        import_batch.refresh_from_db()
        return import_batch

    # ── Legacy synchronous API (kept for backward compat) ─────────────────

    @transaction.atomic
    def process_upload(self, *, batch, uploaded_by, file_name, raw_bytes, original_file_name=None):
        """DEPRECATED: Use the 4-step wizard with process_async instead."""
        if not self.lifecycle.can_edit_records(batch):
            raise ImportRejected(
                'Student records can only be uploaded to a Draft batch.'
            )

        rows = list(self._parse_rows(file_name, raw_bytes))
        if not rows:
            raise ImportRejected('File is empty or unreadable.')

        import_batch = self.batch_repo.create(
            batch=batch, uploaded_by=uploaded_by,
            file_name=file_name or 'upload',
            original_file_name=original_file_name or file_name or 'upload',
            total_rows=len(rows),
            status=ImportBatch.STATUS_PROCESSING,
        )

        success = 0
        skipped = 0
        errors = []

        faculty_by_code = {f.code: f for f in Faculty.objects.all()}
        department_by_code = {d.code: d for d in Department.objects.all()}

        seen_indices = set()
        existing_indices = set(
            StudentRecord.objects
            .filter(batch=batch)
            .values_list('index_number', flat=True)
        )

        for line_no, raw in enumerate(rows, start=2):
            try:
                cleaned = self._clean_row_legacy(raw, faculty_by_code, department_by_code)
            except ValidationError as e:
                errors.append({
                    'row': line_no,
                    'field': getattr(e, 'field', '') or '',
                    'message': '; '.join(e.messages) if hasattr(e, 'messages') else str(e),
                })
                continue

            idx = cleaned['index_number']
            if idx in seen_indices:
                errors.append({'row': line_no, 'field': 'index_number',
                               'message': f"Duplicate index number {idx} in file."})
                continue
            if idx in existing_indices:
                errors.append({'row': line_no, 'field': 'index_number',
                               'message': f"Index number {idx} already exists in this batch."})
                skipped += 1
                continue
            seen_indices.add(idx)

            StudentRecord.objects.create(
                batch=batch, import_batch=import_batch, **cleaned,
            )
            success += 1

        import_batch_status = (
            ImportBatch.STATUS_COMPLETED if not errors
            else ImportBatch.STATUS_COMPLETED_WITH_ERRORS
        )
        ImportBatch.objects.filter(pk=import_batch.pk).update(
            success_count=success,
            skipped_count=skipped,
            error_count=len(errors),
            error_log=errors,
            status=import_batch_status,
            completed_at=timezone.now(),
        )
        import_batch.refresh_from_db()
        return import_batch

    # ── Parsing (raw, preserves original column names) ─────────────────────

    def _parse_rows_raw(self, file_name, raw_bytes) -> Iterable[dict]:
        name = (file_name or '').lower()
        if name.endswith('.xlsx') or name.endswith('.xls'):
            yield from self._parse_xlsx_raw(raw_bytes)
        else:
            yield from self._parse_csv_raw(raw_bytes)

    def _parse_csv_raw(self, raw_bytes):
        try:
            text = raw_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = raw_bytes.decode('latin-1', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            yield {k.strip(): (v or '').strip() for k, v in row.items() if k is not None}

    def _parse_xlsx_raw(self, raw_bytes):
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as e:
            raise ImportRejected('XLSX support requires openpyxl.') from e
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return
        keys = [str(h).strip() if h is not None else '' for h in header]
        for raw_row in rows_iter:
            if not any(c is not None and str(c).strip() for c in raw_row):
                continue
            yield {k: (str(v).strip() if v is not None else '')
                   for k, v in zip(keys, raw_row) if k}

    def _apply_mapping(self, raw: dict, mapping: Dict[str, Optional[str]]) -> dict:
        """Transform raw row (original column names) into system-field dict."""
        result = {}
        for system_field, source_col in mapping.items():
            if source_col:
                result[system_field] = raw.get(source_col, '')
            else:
                result[system_field] = ''
        return result

    # ── Per-row validation & cleaning (mapping-aware) ────────────────────

    def _validate_and_clean_row(self, mapped, faculty_by_code, department_by_code):
        for field, meta in SYSTEM_FIELDS.items():
            if meta['required'] and not mapped.get(field):
                err = ValidationError(f"Missing required value: {meta['label']}")
                err.field = field
                raise err

        email = mapped['institutional_email']
        try:
            validate_email(email)
        except ValidationError as e:
            err = ValidationError(f"Invalid email: {email}")
            err.field = 'institutional_email'
            raise err from e

        # TODO: Re-enable before production — ensure institutional emails
        # if not email.lower().endswith('@uew.edu.gh'):
        #     err = ValidationError(f"Email must be a @uew.edu.gh address: {email}")
        #     err.field = 'institutional_email'
        #     raise err

        completion = self._parse_date(mapped['date_of_completion'])
        if completion is None:
            err = ValidationError(
                f"Invalid date of completion: {mapped['date_of_completion']}"
            )
            err.field = 'date_of_completion'
            raise err

        admission = self._parse_date(mapped.get('date_of_admission')) if mapped.get('date_of_admission') else None

        first_name = mapped.get('first_name', '').strip()
        middle_name = mapped.get('middle_name', '').strip()
        last_name = mapped.get('last_name', '').strip()

        cleaned = {
            'index_number': mapped['index_number'],
            'first_name': first_name,
            'middle_name': middle_name,
            'last_name': last_name,
            'gender': (mapped.get('gender') or '').upper(),
            'institutional_email': email,
            'programme': mapped['programme'],
            'class_of_degree': mapped['class_of_degree'],
            'date_of_completion': completion,
            'date_of_admission': admission,
            'faculty': None,
            'department': None,
            'faculty_name': (mapped.get('faculty') or '').strip(),
            'department_name': (mapped.get('department') or '').strip(),
            'extra_fields': {},
        }

        if cleaned['gender'] and cleaned['gender'] not in {'MALE', 'FEMALE', 'OTHER'}:
            cleaned['gender'] = ''

        class_val = cleaned['class_of_degree']
        if class_val and class_val not in CLASS_OF_DEGREE_CHOICES:
            err = ValidationError(f"Invalid class of degree: {class_val}")
            err.field = 'class_of_degree'
            raise err

        # Faculty / department are free text. We best-effort link to an existing
        # reference entity for reporting, but never reject an unknown value.
        fac_text = cleaned['faculty_name']
        if fac_text:
            faculty = faculty_by_code.get(fac_text) or \
                Faculty.objects.filter(name__iexact=fac_text).first()
            if faculty:
                cleaned['faculty'] = faculty

        dept_text = cleaned['department_name']
        if dept_text:
            department = department_by_code.get(dept_text) or \
                Department.objects.filter(name__iexact=dept_text).first()
            if department:
                cleaned['department'] = department
                if cleaned['faculty'] is None:
                    cleaned['faculty'] = department.faculty
                    cleaned['faculty_name'] = cleaned['faculty_name'] or department.faculty.name

        return cleaned

    # ── Legacy row cleaning (normalised keys) ────────────────────────────

    def _clean_row_legacy(self, raw, faculty_by_code, department_by_code):
        for col in ['index_number', 'first_name', 'last_name', 'institutional_email',
                    'programme', 'class_of_degree', 'date_of_completion']:
            if not raw.get(col):
                err = ValidationError(f"Missing required value: {col}")
                err.field = col
                raise err

        email = raw['institutional_email']
        try:
            validate_email(email)
        except ValidationError as e:
            err = ValidationError(f"Invalid email: {email}")
            err.field = 'institutional_email'
            raise err from e

        completion = self._parse_date(raw['date_of_completion'])
        if completion is None:
            err = ValidationError(f"Invalid date_of_completion: {raw['date_of_completion']}")
            err.field = 'date_of_completion'
            raise err

        admission = self._parse_date(raw.get('date_of_admission')) if raw.get('date_of_admission') else None

        first_name = raw.get('first_name', '').strip()
        middle_name = raw.get('middle_name', '').strip()
        last_name = raw.get('last_name', '').strip()

        cleaned = {
            'index_number': raw['index_number'],
            'first_name': first_name,
            'middle_name': middle_name,
            'last_name': last_name,
            'gender': (raw.get('gender') or '').upper(),
            'institutional_email': email,
            'programme': raw['programme'],
            'class_of_degree': raw['class_of_degree'],
            'date_of_completion': completion,
            'date_of_admission': admission,
            'faculty': None,
            'department': None,
            'faculty_name': (raw.get('faculty_code') or raw.get('faculty') or '').strip(),
            'department_name': (raw.get('department_code') or raw.get('department') or '').strip(),
            'extra_fields': {},
        }

        if cleaned['gender'] and cleaned['gender'] not in {'MALE', 'FEMALE', 'OTHER'}:
            cleaned['gender'] = ''

        # Free-text faculty / department: best-effort FK link, never reject.
        fac_text = cleaned['faculty_name']
        if fac_text:
            faculty = faculty_by_code.get(fac_text) or \
                Faculty.objects.filter(name__iexact=fac_text).first()
            if faculty:
                cleaned['faculty'] = faculty

        dept_text = cleaned['department_name']
        if dept_text:
            department = department_by_code.get(dept_text) or \
                Department.objects.filter(name__iexact=dept_text).first()
            if department:
                cleaned['department'] = department
                if cleaned['faculty'] is None:
                    cleaned['faculty'] = department.faculty
                    cleaned['faculty_name'] = cleaned['faculty_name'] or department.faculty.name

        known = {'index_number', 'first_name', 'middle_name', 'last_name',
                 'institutional_email', 'programme', 'class_of_degree',
                 'date_of_completion', 'gender', 'date_of_admission',
                 'faculty_code', 'department_code'}
        extra = {k: v for k, v in raw.items() if k not in known and k}
        if extra:
            cleaned['extra_fields'] = extra

        return cleaned

    @staticmethod
    def _parse_date(s):
        if not s:
            return None
        if isinstance(s, (date, datetime)):
            return s.date() if isinstance(s, datetime) else s
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

